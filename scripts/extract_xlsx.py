"""엑셀 당직표를 data/schedule.json으로 변환하는 스크립트.

사용법:
    python3 scripts/extract_xlsx.py <엑셀파일경로> [출력경로]

엑셀 파일은 시트 "26년 당직 일정_정리" 안에
  - F4:L.. : 주말/휴일 3팀(영기·소강·도강) 순환 당직 표 (월/날짜/요일/구분/조직/영기 주말 당직/비고)
  - R4:V.. : 영기팀 평일 개별 당직 표 (월/날짜/요일/구분/영기 평일 당직)
구조를 그대로 따라야 합니다. 새 엑셀 파일을 받으면 이 스크립트로 다시 추출한 뒤
data/schedule.json을 갱신하고 커밋하면 됩니다.
"""
import sys
import json
import datetime
import openpyxl

SHEET_NAME = "26년 당직 일정_정리"


def to_date_str(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    return v


def extract(src_path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb[SHEET_NAME]

    weekend = []
    r = 5
    while True:
        date = ws.cell(row=r, column=7).value  # G
        if date is None:
            break
        weekend.append({
            "date": to_date_str(date),
            "weekday": ws.cell(row=r, column=8).value or "",
            "type": ws.cell(row=r, column=9).value or "",
            "org": ws.cell(row=r, column=10).value or "",
            "person": ws.cell(row=r, column=11).value or "",
            "note": ws.cell(row=r, column=12).value or "",
        })
        r += 1

    weekday = []
    r = 5
    while True:
        date = ws.cell(row=r, column=19).value  # S
        if date is None:
            break
        weekday.append({
            "date": to_date_str(date),
            "weekday": ws.cell(row=r, column=20).value or "",
            "type": ws.cell(row=r, column=21).value or "",
            "person": ws.cell(row=r, column=22).value or "",
        })
        r += 1

    return {
        "meta": {
            "weekendRotation": ["영기", "소강", "도강"],
            "source": src_path.split("/")[-1],
            "updatedAt": datetime.datetime.now().isoformat(),
        },
        "weekendDuty": weekend,
        "weekdayDuty": weekday,
    }


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "data/schedule.json"
    data = extract(src)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"주말/휴일 당직 {len(data['weekendDuty'])}건, 평일 당직 {len(data['weekdayDuty'])}건 -> {out}")


if __name__ == "__main__":
    main()
